#!/usr/bin/env python3
"""用 LibreOffice 重算 Excel 公式并报告错误单元格。"""

from __future__ import annotations

import argparse
import json
import os
import platform
import subprocess
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "shared" / "office"))

from soffice import find_soffice, get_soffice_env  # noqa: E402

from openpyxl import load_workbook

ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
MACRO_BODY = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
Sub RecalculateAndSave()
  ThisComponent.calculateAll()
  ThisComponent.store()
  ThisComponent.close(True)
End Sub
</script:module>'''


def macro_dir() -> Path:
    if platform.system() == "Darwin":
        return Path("~/Library/Application Support/LibreOffice/4/user/basic/Standard").expanduser()
    return Path("~/.config/libreoffice/4/user/basic/Standard").expanduser()


def ensure_macro() -> bool:
    d = macro_dir()
    f = d / "Module1.xba"
    if f.is_file() and "RecalculateAndSave" in f.read_text(encoding="utf-8", errors="replace"):
        return True
    d.mkdir(parents=True, exist_ok=True)
    f.write_text(MACRO_BODY, encoding="utf-8")
    return True


def recalc(path: str, timeout: int = 60) -> dict:
    if not Path(path).is_file():
        return {"status": "error", "error": f"文件不存在: {path}"}
    if not find_soffice():
        return {"status": "error", "error": "需要 LibreOffice (soffice)"}
    if not ensure_macro():
        return {"status": "error", "error": "无法配置 LibreOffice 宏"}

    abs_path = str(Path(path).resolve())
    cmd = [
        "soffice",
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]
    if platform.system() == "Linux" and shutil_which("timeout"):
        cmd = ["timeout", str(timeout)] + cmd

    subprocess.run(cmd, capture_output=True, text=True, env=get_soffice_env())

    wb = load_workbook(path, data_only=True)
    summary: dict[str, list[str]] = {e: [] for e in ERRORS}
    total = 0
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str):
                    for err in ERRORS:
                        if err in val:
                            ref = f"{sheet}!{cell.coordinate}"
                            summary[err].append(ref)
                            total += 1
    status = "ok" if total == 0 else "errors_found"
    return {"status": status, "total_errors": total, "error_summary": summary}


def shutil_which(name: str) -> str | None:
    import shutil

    return shutil.which(name)


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("xlsx")
    args = p.parse_args()
    result = recalc(args.xlsx)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result.get("status") == "ok" else 1


if __name__ == "__main__":
    raise SystemExit(main())
